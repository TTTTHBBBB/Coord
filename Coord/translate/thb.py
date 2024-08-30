import openpyxl
import subprocess
from xml.dom.minidom import Document


def read_excel(file_path):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    
    # 读取第一个 sheet 的数据
    sheet = wb.active
    
    # 获取第一行的语言代码和后续行的数据
    languages = []
    rows = []
    for col in range(2, sheet.max_column + 1):
        languages.append(sheet.cell(row=1, column=col).value)
    
    for row in range(2, sheet.max_row + 1):
        rows.append([sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)])
    
    return languages, rows

def create_ts_file(language, keys, translations):
    # 创建 XML 文档
    doc = Document()
    ts = doc.createElement('TS')
    ts.setAttribute('version', '2.1')
    doc.appendChild(ts)
    
    # 创建 <context> 元素
    context = doc.createElement('context')
    ts.appendChild(context)

    for key, translation in zip(keys, translations):
        # 创建 <message> 元素
        message = doc.createElement('message')

        # 创建 <source> 元素
        source_element = doc.createElement('source')
        source_element.appendChild(doc.createTextNode(key))
        message.appendChild(source_element)

        # 创建 <translation> 元素
        translation_element = doc.createElement('translation')
        translation_element.appendChild(doc.createTextNode(translation))
        message.appendChild(translation_element)

        context.appendChild(message)

    # 将 XML 写入文件，设置缩进和换行
    xml_str = doc.toprettyxml(indent="    ", encoding='utf-8').decode('utf-8')
    ts_file_path = f'Language_{language}.ts'
    with open(ts_file_path, 'w', encoding='utf-8') as file:
        file.write(xml_str)
    
    # 调用 lrelease 命令编译 .ts 文件为 .qm 文件
    try:
        subprocess.run([
            r"D:\QT\Tools\QtDesignStudio\qt6_design_studio_reduced_version\bin\lrelease.exe",
            ts_file_path
        ], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error occurred while running lrelease: {e}")

if __name__ == "__main__":
    # 读取 Excel 文件数据
    languages, rows = read_excel("translate.xlsx")
    
    # 假设第一列为键（key），后续列为不同语言的翻译内容
    keys = [row[0] for row in rows]
    translations = [row[1:] for row in rows]

    # 生成每种语言的 .ts 文件
    for language, translation_column in zip(languages, zip(*translations)):
        create_ts_file(language, keys, translation_column)
